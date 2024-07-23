+"""
+Openpyxl is a library for reading and writing Excel 2010 xlsx and xlsm files.
+
+This project provides a function to process a given workbook.
+
+The function takes a filename as input and for each row in the workbook, it calculates the corrected price by multiplying the price by 0.9 and updates the corrected price in a new column.
+
+It also creates a bar chart and adds it to the workbook. The chart displays the corrected prices in the new column.
+
+The function saves the updated workbook with the same filename.
+
+The function currently works for a single workbook, it will not work for thousands of workbooks.
+
+Usage:
+
+from spreadsheet import process_workbook
+process_workbook('filename.xlsx')
+
+"""

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def  process_workbook(filename):

  wb = xl.load_workbook(filename) # making an object of workbook
  sheet =  wb['Sheet1'] # accessing the sheet from workbook
  #cell = sheet['a1']
  #cell = sheet.cell(1,1) # same thing as above line
  # print(cell.value) , print(sheet.max_row)

  for row in range(2, sheet.max_row+ 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9 
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

  value = Reference(sheet,
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4, 
            max_col=4)

  chart = BarChart()
  chart.add_data(value)
  sheet.add_chart(chart, 'e2')

  wb.save(filename)  

  # this for only one spreadsheet, it will not work for thousands of worksheets

